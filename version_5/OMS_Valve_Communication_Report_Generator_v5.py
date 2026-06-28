###############################################################################
# OMS VALVE & COMMUNICATION REPORT GENERATOR
#
# Purpose:
#   Generates Communication Report and Valve Report from OMS communication logs.
#
# Features:
#   - Loads Master Sheet
#   - Loads up to 5 Communication Log files
#   - Removes duplicate packets
#   - Validates packets
#   - Calculates communication statistics
#   - Extracts valve runtime information
#   - Generates formatted Excel reports
#   - Provides Serial Number based Search Mode
#
# Version : 5.0
###############################################################################

import pandas as pd
import os
import re
import time
import sys
from tkinter import Tk, filedialog

###############################################################################
# Extract Valve Information
#
# Reads valve information enclosed inside [] from an OMS packet.
#
# Example:
# [1-1-02:00-01:58-10:00,2-0-01:00-00:30-09:00]
#
# Returns:
#     List of dictionaries containing:
#         valve_no
#         status
#         configured_run_time
#         valve_run_time
#         start_time
###############################################################################

def extract_valves(packet):
    valves = []
    m = re.search(r"\[(.*?)\]", packet)
    if not m:
        return valves

    for item in m.group(1).split(","):
        parts = item.strip().split("-")
        if len(parts) >= 5:
            try:
                valves.append({
                    "valve_no": int(parts[0]),
                    "status": int(parts[1]),
                    "configured_run_time": parts[2],
                    "valve_run_time": parts[3],
                    "start_time": parts[4]
                })
            except Exception:
                continue
    return valves

###############################################################################
# Main Program Loop
#
# Displays the main menu.
# Allows user to:
#     1. Generate Reports
#     2. Exit Program
###############################################################################

while True:

    print("\n" + "=" * 70)
    print("OMS VALVE & COMMUNICATION REPORT GENERATOR")
    print("=" * 70)
    print("1. Select Files and Generate Report")
    print("2. Exit Program")

    main_choice = input("\nEnter Choice : ").strip()

    if main_choice == "2":
        print("Program Closed")
        break

    if main_choice != "1":
        print("Invalid Choice")
        continue
    
    root = None   
    
    try:

        root = Tk()
        root.withdraw()

        print("-" * 70)
        print("SELECT MASTER SHEET")
        print("-" * 70)

        master_file = filedialog.askopenfilename(
            title="Select Master Sheet",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )

        if not master_file:
            print("Master Sheet Not Selected.")
            continue

        print(master_file)

        log_files = []

        print("\n" + "-" * 70)
        print("SELECT COMMUNICATION LOG FILES")
        print("-" * 70)

        while len(log_files) < 5:

            print(f"\nSelect Communication Log File {len(log_files)+1}")

            file = filedialog.askopenfilename(
                title=f"Select Communication Log File {len(log_files)+1}",
                filetypes=[("Excel Files", "*.xlsx *.xls")]
            )

            if file:
                log_files.append(file)
                print(f"Added : {os.path.basename(file)}")

            else:
                print("No file selected.")

            if len(log_files) == 5:
                print("\nMaximum 5 files selected.")
                break

            choice = input(
                "\nPress Enter to select another file or E to start processing: "
            ).strip().upper()

            if choice == "E":
                break

        if len(log_files) == 0:
            print("No Communication Logs Selected.")
            continue
        
        start_time = time.perf_counter()
        
        print("\nLoading Master Sheet...")
        
###############################################################################
# Load Master Sheet
#
# Reads Master Sheet.
# Detects Serial Number and Device ID columns automatically.
# Removes blank records.
# Checks duplicate Serial Numbers and Device IDs.
###############################################################################
        
        master_df = pd.read_excel(master_file)
        master_df.columns = master_df.columns.astype(str).str.strip()

        # Automatically detect required columns
        serial_col = next(
            (col for col in master_df.columns if "SERIAL" in col.upper()),
            None
        )

        device_col = next(
            (col for col in master_df.columns if "DEVICE" in col.upper()),
            None
        )

        if serial_col is None:
            raise Exception("Serial Number column not found.")

        if device_col is None:
            raise Exception("Device ID column not found.")

        # Clean data
        master_df[serial_col] = (
            master_df[serial_col]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        master_df[device_col] = (
            master_df[device_col]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        # Remove blank rows
        master_df = master_df[
            (master_df[serial_col] != "")
            &
            (master_df[device_col] != "")
        ]

        # Duplicate Device IDs
        duplicate_devices = master_df[
            master_df.duplicated(device_col, keep=False)
        ]

        if not duplicate_devices.empty:
            print("\nWARNING : Duplicate Device IDs Found")
            print(duplicate_devices[[serial_col, device_col]])

        # Duplicate Serial Numbers
        duplicate_serials = master_df[
            master_df.duplicated(serial_col, keep=False)
        ]

        if not duplicate_serials.empty:
            print("\nWARNING : Duplicate Serial Numbers Found")
            print(duplicate_serials[[serial_col, device_col]])

        device_to_serial = dict(
            zip(
                master_df[device_col],
                master_df[serial_col]
            )
        )

        serial_to_device = {
            serial: device
            for device, serial in device_to_serial.items()
        }

        print(f"Master Sheet Loaded : {len(device_to_serial)} Devices")

        print("\n" + "=" * 70)
        print("LOADING COMMUNICATION LOG FILES")
        print("=" * 70)

        all_logs = []

        required_columns = [
            "DATE-TIME",
            "MESSAGE"
        ]
        
###############################################################################
# Load Communication Logs
#
# Reads selected communication log files.
# Validates required columns.
# Removes empty MESSAGE records.
# Combines all logs into a single DataFrame.
###############################################################################

        for index, file in enumerate(log_files, start=1):

            print(f"\n[{index}/{len(log_files)}] Reading : {os.path.basename(file)}")

            try:

                df = pd.read_excel(
                    file,
                    header=3
                )
                
                #print(df["DATE-TIME"].head())

                df.columns = (
                    df.columns.astype(str)
                    .str.strip()
                )
                
                # Convert DATE-TIME immediately after reading each file
                df["DATE-TIME"] = pd.to_datetime(
                    df["DATE-TIME"],
                    format="%d-%m-%Y %I:%M:%S %p",
                    errors="coerce"
                )

                # Remove rows with invalid DATE-TIME
                df = df[df["DATE-TIME"].notna()]

                # Check required columns
                missing = [
                    col for col in required_columns
                    if col not in df.columns
                ]

                if missing:

                    print(f"Skipped -> Missing Columns : {missing}")
                    continue

                # Remove empty MESSAGE rows
                df = df[
                    df["MESSAGE"].notna()
                ]

                df = df[
                    df["MESSAGE"].astype(str).str.strip() != ""
                ]

                print(f"Loaded Records : {len(df)}")

                all_logs.append(df)

            except Exception as e:

                print(f"Failed : {os.path.basename(file)}")
                print(f"Reason : {e}")

        if not all_logs:
            raise Exception("No valid communication logs found.")

        log_df = pd.concat(
            all_logs,
            ignore_index=True
        )

        # Remove duplicate packets
        before = len(log_df)
        
###############################################################################
# Remove Duplicate Packets
#
# Duplicate packets are identified using:
#     DATE-TIME
#     MESSAGE
###############################################################################
        
        log_df.drop_duplicates(
            subset=["DATE-TIME", "MESSAGE"],
            inplace=True
        )

        after = len(log_df)

        print("\n" + "-" * 70)
        print(f"Total Records Loaded     : {before}")
        print(f"Duplicate Records Removed: {before - after}")
        print(f"Final Records            : {after}")
        print("-" * 70)

        # Sort packets by time

        log_df = log_df.sort_values(
            "DATE-TIME"
        ).reset_index(drop=True)

        valid_dates = log_df["DATE-TIME"].dropna()

        if valid_dates.empty:
            raise Exception("No valid DATE-TIME found in communication logs.")

        report_date = valid_dates.iloc[0].strftime("%d-%m-%Y")

        print("\n" + "=" * 70)
        print("PROCESSING COMMUNICATION PACKETS")
        print("=" * 70)

        device_summary = {}
        valve_summary = {}

        processed_packets = 0
        invalid_packets = 0
        unknown_devices = 0
        
###############################################################################
# Process Communication Packets
#
# Steps:
#   - Validate packet header
#   - Validate packet length
#   - Verify Device ID
#   - Update Communication Statistics
#   - Extract Valve Information
###############################################################################
        
        for row in log_df.itertuples(index=False):

            packet = str(getattr(row, "MESSAGE", "")).strip()

            if packet == "" or packet.lower() == "nan":
                continue

            # Validate packet header
            if not packet.startswith("##"):
                invalid_packets += 1
                continue

            fields = [x.strip() for x in packet.split(",")]

            # Minimum packet validation
            EXPECTED_MIN_FIELDS = 10

            if len(fields) < EXPECTED_MIN_FIELDS:

                invalid_packets += 1

                continue

            device_id = fields[1].strip().upper()

            if device_id not in device_to_serial:
                unknown_devices += 1
                continue

            serial_no = device_to_serial[device_id]

            processed_packets += 1

###############################################################################
# Communication Summary
#
# Calculates:
#     Total Packets
#     Communication Count
#     No Communication Count
###############################################################################

            stats = device_summary.setdefault(
                device_id,
                {
                    "total": 0,
                    "comm": 0,
                    "no_comm": 0
                }
            )

            stats["total"] += 1

            try:
                comm_status = int(fields[-5])
            except (ValueError, IndexError):
                comm_status = 0

            if comm_status == 1:
                stats["comm"] += 1
            else:
                stats["no_comm"] += 1

###############################################################################
# Valve Processing
#
# Valve information is applicable only for OMS Devices.
# AMS Devices are ignored for valve report generation.
###############################################################################

            if not serial_no.startswith("OMSX"):
                continue

            if device_id not in valve_summary:

                valve_summary[device_id] = {

                    "valves": {

                        valve: {

                            "present": False,
                            "start_time": "",
                            "configured_run_time": "",
                            "valve_run_time": "",
                            "status": "SKIP"

                        }

                        for valve in range(1,9)

                    }

                }

###############################################################################
# Extract Valve Runtime Details
#
# Updates:
#     Start Time
#     Configured Run Time
#     Actual Run Time
#     Valve Status
###############################################################################

            for valve in extract_valves(packet):

                valve_no = valve["valve_no"]

                if valve_no not in range(1,9):
                    continue

                valve_info = valve_summary[device_id]["valves"][valve_no]

                valve_info["present"] = True

                if valve_info["start_time"] == "":
                    valve_info["start_time"] = valve["start_time"]

                valve_info["configured_run_time"] = valve["configured_run_time"]

                valve_info["valve_run_time"] = valve["valve_run_time"]

                config = str(valve["configured_run_time"]).strip()

                runtime = str(valve["valve_run_time"]).strip()

                ##################################################
                # Improved Status Logic
                ##################################################

                if runtime in ("", "00:00", "00:00:00"):

                    valve_info["status"] = "NOT STARTED"

                elif runtime == config:

                    valve_info["status"] = "COMPLETED"

                else:

                    valve_info["status"] = "RUNNING"

        print("\n" + "-" * 70)
        print(f"Processed Packets : {processed_packets}")
        print(f"Unknown Devices   : {unknown_devices}")
        print(f"Invalid Packets   : {invalid_packets}")
        print(f"OMS Devices       : {len(valve_summary)}")
        print(f"Total Devices     : {len(device_summary)}")
        print("-" * 70)

###############################################################################
# Generate Communication Report Data
#
# Calculates:
#     Communication Percentage
#     No Communication Percentage
#
# Separates:
#     AMS Devices
#     OMS Devices
###############################################################################
                
        report_rows = []

        for device_id, stats in device_summary.items():

            total_packets = stats["total"]
            comm_count = stats["comm"]
            no_comm_count = stats["no_comm"]

            comm_pct = round(
                (comm_count / total_packets) * 100, 2
            ) if total_packets else 0

            no_comm_pct = round(
                (no_comm_count / total_packets) * 100, 2
            ) if total_packets else 0

            serial = device_to_serial.get(device_id, "UNKNOWN")

            # Skip UNKNOWN devices
            if serial == "UNKNOWN":
                continue

            report_rows.append({

                "Report Date": report_date,
                "Serial Number": serial,
                "Total Packets": total_packets,
                "Communication Count": comm_count,
                "No Communication Count": no_comm_count,
                "Communication %": comm_pct,
                "No Communication %": no_comm_pct

            })

        comm_df = pd.DataFrame(report_rows)

        # Separate OMS and AMS
        ams_df = (
            comm_df[
                comm_df["Serial Number"].str.startswith("AMSX")
            ]
            .sort_values("Serial Number")
            .reset_index(drop=True)
        )

        oms_df = (
            comm_df[
                comm_df["Serial Number"].str.startswith("OMSX")
            ]
            .sort_values("Serial Number")
            .reset_index(drop=True)
        )

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

###############################################################################
# Create Communication Report Excel
#
# Applies:
#     Header Formatting
#     Borders
#     Auto Column Width
#     Freeze Panes
###############################################################################

        wb = Workbook()
        ws = wb.active
        ws.title = "Communication Report"

        thin = Side(style="thin")

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        header_fill = PatternFill(
            start_color="1F4E79",
            end_color="1F4E79",
            fill_type="solid"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True
        )

        title_font = Font(
            color="FFFFFF",
            bold=True,
            size=13
        )

        center = Alignment(
            horizontal="center",
            vertical="center"
        )

        current_row = 1

###############################################################################
# Write Formatted Table
#
# Writes DataFrame into Excel worksheet with:
#     Title
#     Header Formatting
#     Borders
#     Alignment
###############################################################################

        def write_table(title, df):

            global current_row

            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=len(df.columns)
            )

            cell = ws.cell(
                row=current_row,
                column=1
            )

            cell.value = title
            cell.fill = header_fill
            cell.font = title_font
            cell.alignment = center
            cell.border = border

            current_row += 1

            for col_no, col_name in enumerate(df.columns, 1):

                cell = ws.cell(
                    row=current_row,
                    column=col_no
                )

                cell.value = col_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            current_row += 1

            for row in df.itertuples(index=False):

                for col_no, value in enumerate(row, 1):

                    cell = ws.cell(
                        row=current_row,
                        column=col_no
                    )

                    cell.value = value
                    cell.alignment = center
                    cell.border = border

                current_row += 1

            current_row += 2


        write_table("AMS DEVICES", ams_df)
        write_table("OMS DEVICES", oms_df)

        # Auto adjust column width
        for column_cells in ws.iter_cols(
                min_col=1,
                max_col=ws.max_column):

            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:

                # Ignore merged cells
                if hasattr(cell, "value"):

                    try:
                        if cell.value is not None:
                            max_length = max(
                                max_length,
                                len(str(cell.value))
                            )
                    except:
                        pass

            ws.column_dimensions[column_letter].width = max_length + 3

        comm_file = f"{report_date}_communication.xlsx"
        
        ws.freeze_panes = "A3"

        wb.save(comm_file)

        devices = sorted(
            valve_summary.keys(),
            key=lambda x: device_to_serial.get(x, x)
        )

###############################################################################
# Generate Valve Report
#
# Creates Valve Status Matrix containing:
#     Valve Status
#     Start Time
#     Configured Run Timer
#     Actual Run Timer
###############################################################################

        rows = []

        for valve_no in range(1, 9):

            channel = "CH1" if valve_no <= 4 else "CH2"

            status_row = {
                "CHANNEL": channel,
                "ITEM": f"Valve {valve_no}"
            }

            start_row = {
                "CHANNEL": "",
                "ITEM": "Valve Start Time"
            }

            config_row = {
                "CHANNEL": "",
                "ITEM": "Configured Run Timer"
            }

            run_row = {
                "CHANNEL": "",
                "ITEM": "Valve Run Timer"
            }

            for dev in devices:

                serial = device_to_serial.get(dev, dev)

                valve = valve_summary[dev]["valves"][valve_no]

                if not valve["present"]:

                    status_row[serial] = "SKIP"
                    start_row[serial] = "-"
                    config_row[serial] = "-"
                    run_row[serial] = "-"

                else:

                    status_row[serial] = valve["status"]
                    start_row[serial] = valve["start_time"] or "-"
                    config_row[serial] = valve["configured_run_time"] or "-"
                    run_row[serial] = valve["valve_run_time"] or "-"

            rows.extend([
                status_row,
                start_row,
                config_row,
                run_row
            ])

        valve_df = pd.DataFrame(rows)

        valve_file = f"{report_date}_Valve_Report.xlsx"

###############################################################################
# Format Valve Report Excel
#
# Applies:
#     Header Formatting
#     Borders
#     Cell Coloring
#     Auto Column Width
###############################################################################

        with pd.ExcelWriter(
            valve_file,
            engine="openpyxl"
        ) as writer:

            valve_df.to_excel(
                writer,
                index=False,
                sheet_name="Valve Report"
            )

            wb = writer.book
            ws = writer.sheets["Valve Report"]

            header_fill = PatternFill(
                fill_type="solid",
                start_color="1F4E79"
            )

            header_font = Font(
                bold=True,
                color="FFFFFF"
            )

            thin = Side(style="thin")

            border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            center = Alignment(
                horizontal="center",
                vertical="center"
            )

            # Header formatting
            for cell in ws[1]:

                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center

            # Data formatting
            for row in ws.iter_rows(min_row=2):

                for cell in row:

                    cell.border = border
                    cell.alignment = center

                    if cell.value == "COMPLETED":

                        cell.fill = PatternFill(
                            fill_type="solid",
                            start_color="92D050"
                        )

                    elif cell.value == "RUNNING":

                        cell.fill = PatternFill(
                            fill_type="solid",
                            start_color="FFD966"
                        )

                    elif cell.value == "NOT STARTED":

                        cell.fill = PatternFill(
                            fill_type="solid",
                            start_color="F4B183"
                        )

                    elif cell.value == "SKIP":

                        cell.fill = PatternFill(
                            fill_type="solid",
                            start_color="D9D9D9"
                        )

            # Auto column width
            for column in ws.columns:

                max_length = 0

                letter = get_column_letter(column[0].column)

                for cell in column:

                    try:

                        if cell.value:

                            max_length = max(
                                max_length,
                                len(str(cell.value))
                            )

                    except Exception:
                        pass

                ws.column_dimensions[letter].width = max_length + 3

            ws.freeze_panes = "C2"
        
###############################################################################
# Build Search Dictionaries
#
# Creates in-memory lookup tables for:
#     Communication Report
#     Valve Report
#
# Enables fast Serial Number searching.
###############################################################################

        comm_lookup = {}

        for _, row in comm_df.iterrows():

            serial = str(row["Serial Number"]).strip().upper()

            comm_lookup[serial] = row.to_dict()

        # Valve Report Lookup
        valve_lookup = {}

        for device_id, valve_data in valve_summary.items():

            serial = device_to_serial.get(device_id, "").strip().upper()

            if serial:
                valve_lookup[serial] = valve_data
        
        
        print("\n" + "-" * 70)
        print("REPORT GENERATED SUCCESSFULLY")
        
        print("\nSUMMARY")
        print("-" * 40)
        print(f"Master Devices      : {len(device_to_serial)}")
        print(f"Communication Nodes : {len(device_summary)}")
        print(f"OMS Devices         : {len(oms_df)}")
        print(f"AMS Devices         : {len(ams_df)}")
        print(f"Processed Packets   : {processed_packets}")
        print(f"Unknown Devices     : {unknown_devices}")
        print(f"Invalid Packets     : {invalid_packets}")
        print("-" * 40)
        
        print("-" * 70)

        print(os.path.abspath(comm_file))
        print(os.path.abspath(valve_file))
        
        elapsed_time = time.perf_counter() - start_time

        print("\n" + "-" * 70)
        print(f"Processing Time : {elapsed_time:.2f} Seconds")
        print("-" * 70)

###############################################################################
# Search Mode
#
# Allows user to:
#     View Communication Report
#     View Valve Report
#     Generate New Report
#     Exit Program
###############################################################################

        while True:

            print("\n" + "=" * 90)
            print("SEARCH MODE")
            print("=" * 90)

            print("1. Communication Report")
            print("2. Valve Report")
            print("3. Generate New Report")
            print("4. Exit Program")

            choice = input("\nEnter Choice : ").strip()

###############################################################################
# Communication Report Search
#
# Searches using partial or complete Serial Number.
# Displays Communication Statistics.
###############################################################################

            if choice == "1":

                while True:

                    serial = input(
                        "\nEnter Serial Number (B/3 = Back): "
                    ).strip().upper()

                    if serial in ("3", "B", "BACK"):
                        break

                    matches = sorted(

                        s
                        for s in comm_lookup

                        if serial in s

                    )
                    
                    if not matches:

                        print("\nSerial Number Not Found")
                        continue

                    if len(matches) > 1:

                        print("\nMultiple Devices Found")

                        for s in matches:
                            print(" -", s)

                        continue

                    row = comm_lookup[matches[0]]

                    print("\n" + "-" * 80)
                    print("DEVICE COMMUNICATION REPORT")
                    print("-" * 80)

                    device_id = serial_to_device.get(matches[0], "UNKNOWN")

                    print(f"{'Device ID':<30}: {device_id}")

                    for key, value in row.items():
                        print(f"{key:<30}: {value}")
                    
###############################################################################
# Valve Report Search
#
# Displays valve-wise information including:
#     Status
#     Start Time
#     Configured Run Timer
#     Valve Run Timer
###############################################################################

            elif choice == "2":

                while True:

                    serial = input(
                        "\nEnter Serial Number (B/3 = Back): "
                    ).strip().upper()

                    if serial in ("3", "B", "BACK"):
                        break

                    matches = sorted(

                        s
                        for s in valve_lookup

                        if serial in s

                    )
                    
                    if not matches:

                        print("\nSerial Number Not Found")
                        continue

                    if len(matches) > 1:

                        print("\nMultiple Devices Found")

                        for s in matches:
                            print(" -", s)

                        continue

                    serial = matches[0]

                    if serial.startswith("AMSX"):

                        print("\nValve Report available only for OMS Devices.")
                        continue

                    valve_data = valve_lookup[serial]

                    device_id = serial_to_device.get(serial, "UNKNOWN")

                    print("\n" + "=" * 80)
                    print(f"DEVICE VALVE REPORT : {serial}")
                    print("=" * 80)
                    print(f"Device ID            : {device_id}")
                    print("=" * 80)

                    for valve_no in range(1, 9):

                        valve = valve_data["valves"][valve_no]

                        print(f"\nValve {valve_no}")
                        print("-" * 40)

                        if not valve["present"]:

                            print("Status               : SKIP")
                            print("Valve Start Time     : -")
                            print("Configured Run Timer : -")
                            print("Valve Run Timer      : -")

                        else:

                            print(f"Status               : {valve['status']}")
                            print(f"Valve Start Time     : {valve['start_time']}")
                            print(f"Configured Run Timer : {valve['configured_run_time']}")
                            print(f"Valve Run Timer      : {valve['valve_run_time']}")
                            
###############################################################################
# Return to File Selection
###############################################################################

            elif choice == "3":

                print("\nReturning to File Selection...\n")
                break

###############################################################################
# Exit Program
###############################################################################

            elif choice == "4":

                print("\nProgram Closed Successfully.")
                sys.exit()

            else:

                print("\nInvalid Choice")
            
        continue

###############################################################################
# Global Exception Handler
#
# Displays unexpected runtime errors without terminating the program abruptly.
###############################################################################

    except Exception as e:

        print("\nPROGRAM FAILED")
        print(str(e))
        input("Press Enter To continue...")

###############################################################################
# Cleanup
#
# Releases Tkinter resources before returning to the main menu.
###############################################################################

    finally:

        if root is not None:

            try:
                root.destroy()
            except Exception:
                pass